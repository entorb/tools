#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
reads certain cells from all Excel files found in folder 'input' and creates a report in file out.xlsx
cells to readout are configured in .ini file
"""

# Built-in/Generic Imports
import os
import glob
from configparser import ConfigParser

# Libs
import openpyxl
# see https://openpyxl.readthedocs.io/en/stable/usage.html

# Author and version info
__author__ = "Dr. Torben Menke"
__email__ = "https://entorb.net"
__maintainer__ = __author__
# __copyright__ = "Copyright 2020, My Project"
# __credits__ = ["John", "Jim", "Jack"]
__license__ = "GPL"
__status__ = "Dev"
__version__ = "0.1"

print(f"""Excel-Bulk-Readout
by {__author__}
{__email__}
""")

folder_input = 'input'
file_output = 'out.xlsx'

if os.path.isfile(file_output):
    os.remove(file_output)

config = ConfigParser()
config.read('excel_readout.ini', encoding='utf-8')

# read the config and store the location of the cells to extract into a list
# list containing the cells to extract as tuple: (section, key,value) e.g. ("Sheet1", "A3", "NameForA3")
lSheetsToReadout = []
# dict that holds a list of tuples of (cell_location, cell_title)
dCellsToReadoutPerSheet = {}

l = config.sections()

for sheet in config.sections():
    lSheetsToReadout.append(sheet)  # = sheet name
    lCellsToReadout = []
    for cell_location in config.options(sheet):
        cell_title = config.get(sheet, cell_location)
        # = B9 = Title for Cells B9
        lCellsToReadout.append((cell_location, cell_title))
    dCellsToReadoutPerSheet[sheet] = lCellsToReadout


# create and open output Excel file
workbookOut = openpyxl.Workbook()
sheetOut = workbookOut.active

# write header row
rowOut = 1
colOut = 1
cellOut = sheetOut.cell(row=rowOut, column=colOut)
cellOut.value = "file"
for sheet_name in lSheetsToReadout:
    colOut += 1
    cellOut = sheetOut.cell(row=rowOut, column=colOut)
    cellOut.value = 'sheet'
    for (cell_location, cell_title) in dCellsToReadoutPerSheet[sheet_name]:
        colOut += 1
        cellOut = sheetOut.cell(row=rowOut, column=colOut)
        cellOut.value = cell_title


# loop over all Excel files in dir "input"
for filepath in glob.glob(folder_input + "/" + "*.xls*"):
    (dirName, fileName) = os.path.split(filepath)

    # open Excel file
    print(f"reading {fileName}")
    # data_only : read values instead of formulas
    workbook = openpyxl.load_workbook(filepath, data_only=True)

    rowOut += 1
    colOut = 1
    cellOut = sheetOut.cell(row=rowOut, column=colOut)
    cellOut.value = fileName

    for sheet_name in lSheetsToReadout:
        colOut += 1
        cellOut = sheetOut.cell(row=rowOut, column=colOut)
        cellOut.value = sheet_name
        for (cell_location, cell_title) in dCellsToReadoutPerSheet[sheet_name]:
            colOut += 1
            sheetIn = workbook[sheet_name]
            cellIn = sheetIn[cell_location]
            cellOut = sheetOut.cell(row=rowOut, column=colOut)
            cellOut.value = cellIn.value

workbookOut.save(file_output)
input("press Enter to close")
