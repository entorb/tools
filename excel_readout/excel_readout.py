#!/usr/bin/env python3


"""
Read many Excels and write into single report Excel.

* read all Excel files in dir `input`
* extract certain cells and
* write into single report Excel `out.xlsx`

configure in `excel_readout.ini`

see README.md
"""

import os
from configparser import ConfigParser
from pathlib import Path

import openpyxl

dir_input = "input"
file_output = Path("out.xlsx")

if file_output.exists():
    file_output.unlink()

config = ConfigParser()
config.read("excel_readout.ini", encoding="utf-8")

# read the config and store the location of the cells to extract into a list
# list containing the cells to extract as tuple: (section, key,value) e.g.
# ("Sheet1", "A3", "NameForA3")
l_sheets_to_read: list[str] = []
# dict that holds a list of tuples of (cell_location, cell_title)
d_cells_to_read_per_sheet = {}

list_of_sections = config.sections()

for sheet in config.sections():
    l_sheets_to_read.append(sheet)  # = sheet name
    l_cells_to_read: list[tuple[str, str]] = []
    for _cell_location in config.options(sheet):
        cell_title = config.get(sheet, _cell_location)
        # = B9 = Title for Cells B9
        l_cells_to_read.append((_cell_location, cell_title))
    d_cells_to_read_per_sheet[sheet] = l_cells_to_read


# create and open output Excel file
workbook_out = openpyxl.Workbook()
sheet_out = workbook_out.active
assert sheet_out  # noqa: S101

# write header row
row_out = 1
col_out = 1
cell_out = sheet_out.cell(row=row_out, column=col_out)
cell_out.value = "file"
for sheet_name in l_sheets_to_read:
    col_out += 1
    cell_out = sheet_out.cell(row=row_out, column=col_out)
    cell_out.value = "sheet"
    for _cell_location, cell_title in d_cells_to_read_per_sheet[sheet_name]:  # type: ignore
        col_out += 1
        cell_out = sheet_out.cell(row=row_out, column=col_out)
        cell_out.value = cell_title


# loop over all Excel files in dir "input"
for filepath in Path(dir_input).glob("*.xls*"):
    (dir_name, file_name) = os.path.split(filepath)

    # open Excel file
    print(f"reading {file_name}")
    # data_only : read values instead of formulas
    workbook = openpyxl.load_workbook(filepath, data_only=True)

    row_out += 1
    col_out = 1
    cell_out = sheet_out.cell(row=row_out, column=col_out)
    cell_out.value = file_name

    for sheet_name in l_sheets_to_read:
        col_out += 1
        cell_out = sheet_out.cell(row=row_out, column=col_out)
        cell_out.value = sheet_name
        for cell_location, _cell_title in d_cells_to_read_per_sheet[sheet_name]:  # type: ignore
            col_out += 1
            sheet_in = workbook[sheet_name]
            cell_in = sheet_in[cell_location]
            cell_out = sheet_out.cell(row=row_out, column=col_out)
            cell_out.value = cell_in.value  # type: ignore

workbook_out.save(file_output)
input("press Enter to close")
